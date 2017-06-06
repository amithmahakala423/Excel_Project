import openpyxl
from configparser import SafeConfigParser
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Series, Reference
import datetime
import time
from datetime import date, timedelta



class ExcelProject(object):
#########################################################################
    def __init__(self):

        self.start_time = time.time()
        #initializing the configparser
        excel_parser = SafeConfigParser()
        excel_parser.read('excel_project.ini')

        #creating global_values
        self.number = int(excel_parser.get('inflow', 'number'))
        self.number_toassign = int(excel_parser.get('inflow', 'number'))
        self.app_cell = str(excel_parser.get('inflow', 'app_cell'))
        self.app_cell_toassign = str(excel_parser.get('inflow', 'app_cell'))
        self.date_cell = str(excel_parser.get('inflow', 'date_cell'))
        self.date_cell_toassign = str(excel_parser.get('inflow', 'date_cell'))
        self.repeat_toassign = int(excel_parser.get('inflow', 'repeat'))

        self.number_aging = int(excel_parser.get('aging', 'number_aging'))
        self.number_aging_toassign = int(excel_parser.get('aging', 'number_aging'))
        self.app_aging_cell = str(excel_parser.get('aging', 'app_aging_cell'))
        self.app_aging_cell_toassign = str(excel_parser.get('aging', 'app_aging_cell'))
        self.status_aging_cell = str(excel_parser.get('aging', 'status_aging_cell'))
        self.status_aging_cell_toassign = str(excel_parser.get('aging', 'status_aging_cell'))
        self.label_aging_cell = str(excel_parser.get('aging', 'label_aging_cell'))
        self.label_aging_cell_toassign = str(excel_parser.get('aging', 'label_aging_cell'))
        
        self.count = 1
        #creating a empty dictionary
        self.app_dict = {}
        #assigning values by reading from ini file
        self.field_pre = 'app'
        self.field_suff = self.field_pre + str(self.count)
        
        while str(excel_parser.get('app_dict', self.field_suff)) != '':
            self.app_dict[self.count] = str(excel_parser.get('app_dict', self.field_suff))
            self.count += 1
            self.field_suff = self.field_pre + str(self.count)

        print ("total no of apps: "+ str(self.count -1))
        print ("final dict is: "+ str(self.app_dict))
        print ()
        print ()

        
        #target sheet initialization
        print ("Initialising the target sheet")
        self.graph_workbook = Workbook(write_only= True)
        self.graph_workbook_sheet = self.graph_workbook.create_sheet()
        print ("Default sheet name is: "+ str(self.graph_workbook_sheet))
        self.graph_workbook_sheet1 = self.graph_workbook.create_sheet()
        print ("Second sheet name is: "+ str(self.graph_workbook_sheet1))
        print ()
        print ()

        
        #source sheet initialization
        print ("Loading the workbook")
        print ("Workbook name: " + str(excel_parser.get('workbook', 'name')))
        self.report_workbook = load_workbook(str(excel_parser.get('workbook', 'name')))
        #getting the active sheet in source
        self.report_workbook_sheet = self.report_workbook.active
        print ("Work sheet name is:"+ str(self.report_workbook_sheet))
        print ()
        print ()

        #initialize the list and append the first row.
        self.rows =[]
        for i in range(self.count):
            self.rows.append([])
        #pushing the first row to the list
        self.rows[0].append('Row Labels')
        self.nth = 7
        while self.nth > 0:
            self.rows[0].append(date.today() - timedelta(self.nth))
            self.nth -= 1


        #initializing the list and append the first row for aging values
        self.rows_aging =[]
        for i in range(self.count):
            self.rows_aging.append([])
        #pushing the first row to the list
        self.rows_aging[0].append('Row Labels')
        self.rows_aging[0].append('> 90 Days')
        self.rows_aging[0].append('22 - 90 Days')
        self.rows_aging[0].append('8 - 21 Days')
        self.rows_aging[0].append('0 - 7 Days')
#######################################################################
    #function for getting cell values
    def cell_value(self,number):
        self.app_cell = 'B' + str(self.number)
        self.date_cell = 'F' + str(self.number)


    def access_incident_values(self):
        for i in range(1,self.count):
            self.application_name = self.app_dict[i]
            #print ("checking for application: "+ self.application_name)
            self.rows[i].append(self.application_name)
            self.repeat = self.repeat_toassign
            self.application_incident_count = 0
            print ("Starting Application: " + self.application_name)
            while self.repeat > 0 :
                #print ("self.repeat value is:" + str(self.repeat))
                self.full_date = str(date.today() - timedelta(self.repeat))
                self.compare_date = self.full_date[:10]
                #print ("self.compare_date becomes: "+ self.compare_date)
                #print ("date_cell renamed to: " + str(self.date_cell))
                while True:
                    #selecting the application name and counting the IRs on the particular
                    #day, and then pushing it to the list
                    
                    self.time_stamp = str(self.report_workbook_sheet[self.date_cell].value)
                    self.actual_time = self.time_stamp[:10]
                    #print("actual time: " + self.actual_time)
                    if self.compare_date == self.actual_time and self.application_name == str(self.report_workbook_sheet[self.app_cell].value):
                        self.application_incident_count += 1
                    self.number += 1
                    self.cell_value(self.number)
                    #print (self.app_cell)
                    #print (self.date_cell)
                    check_none = self.report_workbook_sheet[self.app_cell].value
                    #print ("check_none: "+ str(check_none))
                    if str(check_none) == "None":
                        self.rows[i].append(self.application_incident_count)
                        #print ("self.app_name" + self.application_name + "date: " + self.compare_date + \
                               #"self.application_incident_count "+ str(self.application_incident_count ))
                        break
                    
                self.repeat -= 1
                self.application_incident_count = 0
                self.date_cell = self.date_cell_toassign
                self.app_cell = self.app_cell_toassign
                self.number = self.number_toassign
            print ("Finished Application: " + self.application_name)
            print ()
            print ()
                
        
        print (self.rows)



    def append_incident_values_to_sheet(self):
        for row_values in self.rows:
            self.graph_workbook_sheet.append(row_values)
        print ("Alvalues got appended to sheet")



    def draw_incident_inflow_graph(self):
        chart = BarChart()
        chart.type = "col"
        chart.title = "IR Inflow - Last 7 Days - Top 10 Apps"
        chart.style = 10
        chart.x_axis.title = 'Applications'
        chart.y_axis.title = 'Inflow_Value'
        data = Reference(self.graph_workbook_sheet, min_col=2, min_row=1, max_row= 6, max_col=8)
        cats = Reference(self.graph_workbook_sheet, min_col=1, min_row=2, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 6
        self.graph_workbook_sheet.add_chart(chart, "K2")    
        #self.graph_workbook.save('final_graph.xlsx')
        #print ("######SUCCESSFUL......GRAPH IS READY######")
        
        
#######################################################################

    def access_aging_values(self):
        for i in range(1,self.count):
            self.application_name = self.app_dict[i]
            #print ("checking for application: "+ self.application_name)
            self.rows_aging[i].append(self.application_name)
            self.repeat_aging = 1
            self.application_aging_count = 0
            print ("Starting Application: " + self.application_name)
            while self.repeat_aging < 5 :
                #print ("self.repeat_aging value is:" + str(self.repeat_aging))
                #print ("checking for age" + self.rows_aging[0][self.repeat_aging])
                while True:
                    #selecting the application name and counting the IRs on the particular
                    #aging value, and then pushing it to the list
                    self.aging_value = str(self.report_workbook_sheet[self.label_aging_cell].value)
                    self.status = str(self.report_workbook_sheet[self.status_aging_cell].value)
                    if self.aging_value == self.rows_aging[0][self.repeat_aging]  and (self.status != 'CLOSED' \
                                                                                      and self.status != 'COMPLETED'\
                                                                                      and self.status != 'PENDING') \
                                                                                      and self.application_name ==\
                                                                                      self.report_workbook_sheet[self.app_aging_cell].value :
                        self.application_aging_count += 1
                    self.number_aging += 1
                    self.cell_value_aging(self.number_aging)
                    #print (self.app_cell)
                    #print (self.date_cell)
                    check_none = self.report_workbook_sheet[self.app_aging_cell].value
                    #print ("check_none: "+ str(check_none))
                    if str(check_none) == "None":
                        self.rows_aging[i].append(self.application_aging_count)
                        #print ("self.app_name" + self.application_name + "status " + self.status + \
                        #       "self.application_incident_count "+ str(self.application_aging_count ))
                        break
                    
                self.repeat_aging += 1
                self.application_aging_count = 0
                self.status_aging_cell = self.status_aging_cell_toassign
                self.app_aging_cell = self.app_aging_cell_toassign
                self.label_aging_cell = self.label_aging_cell_toassign
                self.number_aging = 2
            print ("Finished Application: " + self.application_name)
            print ()
            print ()
                
        
        print (self.rows_aging)


    #function for getting cell values
    def cell_value_aging(self,number):
        self.app_aging_cell = 'B' + str(self.number_aging)
        self.status_aging_cell = 'H' + str(self.number_aging)
        self.label_aging_cell = 'X' + str(self.number_aging)


    def append_aging_values_to_sheet(self):
        for row_values in self.rows_aging:
            self.graph_workbook_sheet1.append(row_values)
        print ("Alvalues got appended to sheet1")

    def draw_aging_inflow_graph(self):
        chart = BarChart()
        chart.type = "col"
        chart.title = "IR Aging - Top 5 Apps"
        chart.style = 10
        chart.x_axis.title = 'Applications'
        chart.y_axis.title = 'Aging value'
        data = Reference(self.graph_workbook_sheet1, min_col=2, min_row=1, max_row= 6, max_col=5)
        cats = Reference(self.graph_workbook_sheet1, min_col=1, min_row=2, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 6
        self.graph_workbook_sheet1.add_chart(chart, "G2")    
        self.graph_workbook.save('final_graph.xlsx')
        print ("######SUCCESSFUL......GRAPH IS READY######")
        print ()
        print ()
        print("--- %s seconds ---" % (time.time() - self.start_time))

####################################################################
#Methods to read the incident values and draw a graph for the inflow
excel_proj = ExcelProject()
excel_proj.access_incident_values()
excel_proj.append_incident_values_to_sheet()
excel_proj.draw_incident_inflow_graph()

#methods to read the values and draw a graph for the aging of the IRs.
excel_proj.access_aging_values()
excel_proj.append_aging_values_to_sheet()
excel_proj.draw_aging_inflow_graph()



