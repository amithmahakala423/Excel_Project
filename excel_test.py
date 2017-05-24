import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Series, Reference

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

time_stamp= ws['A2']
time_list = str(time_stamp.value)
actual_time = time_list[:10]
if actual_time == '2017-05-22':
    print (actual_time)

time_stamp1= ws['A6']
print (time_stamp1.value)
if str(time_stamp1.value) == "None":
    print ("empty cell")

# Save the file
wb.save("sample.xlsx")
