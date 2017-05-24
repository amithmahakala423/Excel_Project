import openpyxl
from openpyxl import Workbook, load_workbook

class ExcelProject(object):

    def __init__(self):    
        #creating global_values
        self.number = 1
        self.app_cell = 'B2'
        self.date_cell = 'F2'
        self.count = 0

"""1 : 'BIS',
                            2 : 'BPMS',
                            3 : 'Business Data Solutions - MDB',
                            4 : 'Client Portal' ,
                            5 : 'Client Portal Mobile - B2B',
                            6 : 'Content Assistant' ,
                            7 : 'DexBis' ,
                            8 : 'Dexnet' ,
                            9 : 'Distribution' ,
                            10 : 'DSS Request' ,
                            11 : 'Enterprise Apps - kGen' ,
                            12 : 'Enterprise Service Bus' ,
                            13 : 'Intranet' ,
                            14 : 'LSAP & PRP' ,
                            15 : 'Merchant Platform' ,
                            16 : 'NEMO' ,
                            17 : 'Proposal Builder' ,
                            18 : 'Salesforce Platform' ,
                            19 : 'SEM Estimator Tool' ,
                            20 : 'Superpages.com' ,
                            22 : 'Vision' """
        #target sheet initialization
        self.graph_workbook = Workbook(write_only= True)
        self.graph_workbook_sheet = self.graph_workbook.create_sheet()
        print ("Default sheet name is: "+ str(self.graph_workbook_sheet))
        #source sheet initialization
        self.report_workbook = load_workbook('report1495463480142.xlsx')
        #getting the active sheet in source
        self.report_workbook_sheet = self.report_workbook.active
        print ("Work sheet name is:"+ str(self.report_workbook_sheet))



    #function for getting cell values
    def cell_value(self,number):
        self.app_cell = 'B' + str(number)
        self.date_cell = 'F' + str(number)


    #get all the values from the sheet
    def access_values(self): 
        while True:
            time_stamp = str(self.report_workbook_sheet[self.date_cell].value)
            actual_time = time_stamp[:10]
            if actual_time == '2017-05-16' or actual_time == '2017-05-16' or actual_time == '2017-05-18' or \
            actual_time == '2017-05-19' or actual_time == '2017-05-20' or actual_time == '2017-05-21' or actual_time == '2017-05-22':
                self.count += 1
                print (self.report_workbook_sheet[self.app_cell].value, end =" ")
                print (self.report_workbook_sheet[self.date_cell].value)

            self.cell_value(self.number)
            self.number += 1
            check_none = self.report_workbook_sheet[self.app_cell].value
            if str(check_none) == "None":
                print (self.count)
                break



excel_proj = ExcelProject()
excel_proj.access_values()
