import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Series, Reference

### this is the target sheet ###
print ("configuring the target sheet")
graph_workbook = Workbook(write_only=True)
graph_sheet = graph_workbook.create_sheet()
print ("Default sheetname is: " +  str(graph_sheet))
print ("Renaming the sheet")
#graph_sheet.title = "Inflow_Graph"
print ("New sheetname is: "+ str(graph_sheet))


###source sheet configuration ###
report_workbook = load_workbook('Daily Status Report-05082017.xlsx')

#getting required sheet object
ws_IR_Inflow_Chart = report_workbook['IR Inflow Chart']
print ("i am accessing the following sheet in source: " + str(ws_IR_Inflow_Chart))

rows =[]
for i in range(22):
    rows.append([])
#iterating over rows
i =-1

for row1 in ws_IR_Inflow_Chart.iter_rows( min_row = 1, max_row=22, max_col= 8):
    i += 1
    for cell1 in row1:
        print (cell1.value, end = " ")
        rows[i].append(cell1.value)
    print ("")


print("Now designing the graph for the data")
for row_values in rows:
    graph_sheet.append(row_values)


chart = BarChart()
chart.type = "col"
chart.title = "IR Inflow - Last 7 Days - Top 5 Apps"
chart.style = 10
chart.x_axis.title = 'Applications'
chart.y_axis.title = 'Inflow_Value'

data = Reference(graph_sheet, min_col=2, min_row=1, max_row= 6, max_col=8)
cats = Reference(graph_sheet, min_col=1, min_row=2, max_row=6)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4

graph_sheet.add_chart(chart, "K2")    

graph_workbook.save('graph1.xlsx')
